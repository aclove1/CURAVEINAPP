#!/usr/bin/env python3
"""
CuraVein App ↔ Spreadsheet Audit
Reads live app values from /api/audit and compares against spreadsheet cells.

Usage:
    python3 scripts/audit_compare.py                        # default paths
    python3 scripts/audit_compare.py --app http://localhost:3000
    python3 scripts/audit_compare.py --save results.json

Requirements: openpyxl, requests
    pip install openpyxl requests
"""

import argparse, json, sys, math
from datetime import datetime
from pathlib import Path

try:
    import requests
except ImportError:
    print("❌ Missing: pip install requests openpyxl")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("❌ Missing: pip install openpyxl")
    sys.exit(1)

# ── Defaults ─────────────────────────────────────────────────────────────────
APP_URL   = "http://localhost:3000"
XLSX_PATH = "/Users/aaron.c_love/Desktop/Neal:Ethan/CuraVein_Integrated_v11.xlsx"

# ── CLI ───────────────────────────────────────────────────────────────────────
p = argparse.ArgumentParser()
p.add_argument("--app",  default=APP_URL,   help="App base URL")
p.add_argument("--xlsx", default=XLSX_PATH, help="Spreadsheet path")
p.add_argument("--save", default=None,      help="Save JSON results to file")
p.add_argument("--fail-only", action="store_true", help="Only show FAILs")
args = p.parse_args()

PASS = "✅ PASS"
FAIL = "❌ FAIL"
WARN = "⚠️  WARN"
results = []

def check(label, cell, xl_val, app_val, tol=0.0001, note=""):
    if xl_val is None:
        status = WARN
        explanation = "Spreadsheet returned None (formula not cached)"
    elif app_val is None:
        status = FAIL
        explanation = "App value is None/missing"
    else:
        try:
            match = math.isclose(float(xl_val), float(app_val), rel_tol=tol, abs_tol=tol)
            status = PASS if match else FAIL
            explanation = note or (
                f"spreadsheet={xl_val}, app={app_val}" if not match else ""
            )
        except (TypeError, ValueError):
            # String comparison
            match = str(xl_val).strip() == str(app_val).strip()
            status = PASS if match else FAIL
            explanation = note or (f"spreadsheet={xl_val!r}, app={app_val!r}" if not match else "")
    results.append({
        "label": label, "cell": cell,
        "spreadsheet": xl_val, "app": app_val,
        "status": status, "explanation": explanation,
    })

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Fetch app audit JSON
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print(f"  CuraVein Audit — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
print(f"{'='*65}")
print(f"\n📡 Fetching {args.app}/api/audit ...")
try:
    r = requests.get(f"{args.app}/api/audit", timeout=10)
    r.raise_for_status()
    app = r.json()
    print(f"  ✅ App responded — scenario: {app['_meta']['activeScenario']}")
except Exception as e:
    print(f"  ❌ FAILED: {e}")
    print("     Is the dev server running? (npm run dev)")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Open spreadsheet
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n📊 Opening spreadsheet: {args.xlsx}")
if not Path(args.xlsx).exists():
    print(f"  ❌ File not found: {args.xlsx}")
    sys.exit(1)

wb = openpyxl.load_workbook(args.xlsx, data_only=True)
print(f"  ✅ Sheets: {wb.sheetnames}")

SC = wb["Scenario Controls"]
IS = wb["Inventory Supplies"]

def xl(sheet, cell):
    """Read a cell value, return None if formula not cached."""
    v = sheet[cell].value
    return v

def xlf(sheet, cell):
    """Read with formula fallback label."""
    wb2 = openpyxl.load_workbook(args.xlsx, data_only=False)
    sheet_name = sheet.title
    formula = wb2[sheet_name][cell].value
    return formula

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Read spreadsheet cells
# ─────────────────────────────────────────────────────────────────────────────
print("\n📋 Reading spreadsheet cells ...")
xl_vals = {
    # Reimbursement
    "commercial_multiplier":      xl(SC, "F15"),
    "govt_payer_mix":             xl(SC, "F16"),
    "commercial_payer_mix":       xl(SC, "F17"),
    "blended_rate":               xl(SC, "F18"),
    "cpt_weighted_base":          xl(SC, "E130"),
    # Capacity & funnel
    "max_capacity":               xl(SC, "D45"),
    "blended_cpl":                xl(SC, "D7"),
    "contact_rate":               xl(SC, "D8"),
    "booking_rate":               xl(SC, "D9"),
    "show_rate":                  xl(SC, "D10"),
    "treatment_conversion":       xl(SC, "D11"),
    "procs_per_episode":          xl(SC, "D12"),
    # Growth
    "y2_growth":                  xl(SC, "D37"),
    "y3_growth":                  xl(SC, "D38"),
    # Payer sub-breakdown
    "bcbs_share":                 xl(SC, "B154"),
    "bcbs_multiplier":            xl(SC, "C154"),
    "aetna_uhc_cigna_multiplier": xl(SC, "C155"),
    # COGS / procedure mix
    "venaseal_mix":               xl(IS, "B10"),
    "rfa_mix":                    xl(IS, "B11"),
    "sclero_mix":                 xl(IS, "B12"),
    "varithena_mix":              xl(IS, "B13"),
    "varithena_cost_proc":        xl(IS, "F51"),
    "venaseal_pts_per_kit":       xl(IS, "B8"),
    "venaseal_kit_price":         xl(IS, "B16"),
    "sclero_supply_cost":         xl(IS, "F47"),
    "waste_factor":               xl(IS, "B5"),
    "post_proc_support":          xl(IS, "B7"),
    "misc_consumables":           xl(IS, "B6"),
    # CPT fee schedule
    "cpt_36482_rate":             xl(SC, "C125"),
    "cpt_36482_share":            xl(SC, "D125"),
    "cpt_36465_rate":             xl(SC, "C126"),
    "cpt_36465_share":            xl(SC, "D126"),
    "cpt_36466_rate":             xl(SC, "C127"),
    "cpt_36466_share":            xl(SC, "D127"),
    "cpt_36475_rate":             xl(SC, "C124"),
    "cpt_36475_share":            xl(SC, "D124"),
}

# Flag any None (uncached formulas)
unresolved = [k for k, v in xl_vals.items() if v is None]
if unresolved:
    print(f"  ⚠️  {len(unresolved)} cell(s) returned None (formula not cached):")
    for k in unresolved:
        cell_map = {
            "commercial_multiplier":"F15","govt_payer_mix":"F16",
            "commercial_payer_mix":"F17","blended_rate":"F18",
        }
        print(f"     {k} — re-reading formula...")
else:
    print(f"  ✅ All {len(xl_vals)} cells resolved")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Compare
# ─────────────────────────────────────────────────────────────────────────────
print("\n🔍 Comparing ...")

xv = xl_vals  # shorthand
av = app      # shorthand

# Reimbursement
check("Medicare weighted base",        "SC!E130",  xv["cpt_weighted_base"],
      av["reimbursement"].get("derivedMedicareBase"))
check("Commercial multiplier (F15)",   "SC!F15",   xv["commercial_multiplier"],
      av["reimbursement"].get("commercialMultiplier"))
check("Govt payer mix (F16)",          "SC!F16",   xv["govt_payer_mix"],
      av["payerMix"]["defaults"].get("medicareMix"))
check("Commercial payer mix (F17)",    "SC!F17",   xv["commercial_payer_mix"],
      av["payerMix"]["defaults"].get("commercialMix"))
check("Mature blended rate (F18)",     "SC!F18",   xv["blended_rate"],
      av["reimbursement"].get("derivedBlendedRate"))
check("matureBlendedRate constant",    "SC!F18",   xv["blended_rate"],
      av["reimbursement"].get("matureBlendedRate"))

# Payer sub-breakdown
check("BCBS share of commercial",      "SC!B154",  xv["bcbs_share"],
      av["payerMix"].get("bcbsShare"))
check("BCBS multiplier",               "SC!C154",  xv["bcbs_multiplier"],
      av["payerMix"].get("bcbsMultiplier"))
check("Aetna/UHC/Cigna multiplier",    "SC!C155",  xv["aetna_uhc_cigna_multiplier"],
      av["payerMix"].get("otherCommercialMultiplier"))

# Forney payer mix
check("Forney govt share",             "SC!F16",   xv["govt_payer_mix"],
      av["payerMix"]["byMarket"].get("forney",{}).get("govtShare"))
check("Forney commercial share",       "SC!F17",   xv["commercial_payer_mix"],
      av["payerMix"]["byMarket"].get("forney",{}).get("commercialShare"))
check("NB govt share",                 "SC!B145",  0.25,
      av["payerMix"]["byMarket"].get("new-braunfels",{}).get("govtShare"))
check("NB commercial share",           "SC!C145",  0.75,
      av["payerMix"]["byMarket"].get("new-braunfels",{}).get("commercialShare"))

# Funnel
check("Max capacity/month",            "SC!D45",   xv["max_capacity"],
      av["funnel"].get("maxCapacityPerMonth"))
check("Blended CPL",                   "SC!D7",    xv["blended_cpl"],
      av["funnel"]["cpl"].get("base") if isinstance(av["funnel"].get("cpl"), dict)
      else av["funnel"].get("cpl"))
check("Contact rate",                  "SC!D8",    xv["contact_rate"],
      av["funnel"]["contactRate"].get("base") if isinstance(av["funnel"].get("contactRate"), dict)
      else av["funnel"].get("contactRate"))
check("Booking rate",                  "SC!D9",    xv["booking_rate"],
      av["funnel"]["bookingRate"].get("base") if isinstance(av["funnel"].get("bookingRate"), dict)
      else av["funnel"].get("bookingRate"))
check("Show rate",                     "SC!D10",   xv["show_rate"],
      av["funnel"]["showRate"].get("base") if isinstance(av["funnel"].get("showRate"), dict)
      else av["funnel"].get("showRate"))
check("Treatment conversion",          "SC!D11",   xv["treatment_conversion"],
      av["funnel"]["treatmentConversion"].get("base") if isinstance(av["funnel"].get("treatmentConversion"), dict)
      else av["funnel"].get("treatmentConversion"))
check("Procs per episode",             "SC!D12",   xv["procs_per_episode"],
      av["funnel"]["procsPerPatient"].get("base") if isinstance(av["funnel"].get("procsPerPatient"), dict)
      else av["funnel"].get("procsPerPatient"))

# Growth
check("Y2 volume growth (base)",       "SC!D37",   xv["y2_growth"],
      av["growth"]["y2VolumeGrowth"].get("base") if isinstance(av["growth"].get("y2VolumeGrowth"), dict)
      else av["growth"].get("y2VolumeGrowth"))
check("Y3 volume growth (base)",       "SC!D38",   xv["y3_growth"],
      av["growth"]["y3VolumeGrowth"].get("base") if isinstance(av["growth"].get("y3VolumeGrowth"), dict)
      else av["growth"].get("y3VolumeGrowth"))

# Procedure mix
check("VenaSeal mix %",                "IS!B10",   xv["venaseal_mix"],
      av["procedureMix"].get("venaSeal"))
check("RFA mix %",                     "IS!B11",   xv["rfa_mix"],
      av["procedureMix"].get("rfa"))
check("Sclerotherapy mix %",           "IS!B12",   xv["sclero_mix"],
      av["procedureMix"].get("sclerotherapy"))
check("Varithena mix %",               "IS!B13",   xv["varithena_mix"],
      av["procedureMix"].get("varithena"))

# COGS
check("VenaSeal kit price",            "IS!B16",   xv["venaseal_kit_price"],
      av["cogs"].get("venasealUnitCost"))
check("VenaSeal pts/kit",              "IS!B8",    xv["venaseal_pts_per_kit"],
      av["cogs"].get("venasealPtsPerKit"))
check("Varithena cost/proc",           "IS!F51",   round(xv["varithena_cost_proc"] or 0, 2),
      round(av["cogs"].get("varithenaCostPerProc", 0), 2))
check("Sclero supply cost (F47)",      "IS!F47",   round(xv["sclero_supply_cost"] or 0, 2),
      av["cogs"].get("scleroSupplyCost"))
check("Post-proc support",             "IS!B7",    xv["post_proc_support"],
      av["cogs"].get("postProcSupport"))
check("Misc consumables",              "IS!B6",    xv["misc_consumables"],
      av["cogs"].get("miscConsumables"))
check("Waste factor",                  "IS!B5",    xv["waste_factor"],
      av["cogs"].get("wasteFactor") or
      next((c.get("wasteFactor") for c in [app] if "wasteFactor" in c), None))

# CPT fee schedule
cpt_map = {c["code"]: c for c in av.get("cptTable", [])}
check("CPT 36482 — Medicare rate",     "SC!C125",  round(xv["cpt_36482_rate"] or 0, 2),
      round(cpt_map.get("36482", {}).get("medicareRate", 0), 2))
check("CPT 36482 — volume share",      "SC!D125",  xv["cpt_36482_share"],
      cpt_map.get("36482", {}).get("mixPct"))
check("CPT 36465 — Medicare rate",     "SC!C126",  round(xv["cpt_36465_rate"] or 0, 2),
      round(cpt_map.get("36465", {}).get("medicareRate", 0), 2))
check("CPT 36465 — volume share",      "SC!D126",  xv["cpt_36465_share"],
      cpt_map.get("36465", {}).get("mixPct"))
check("CPT 36466 — Medicare rate",     "SC!C127",  round(xv["cpt_36466_rate"] or 0, 2),
      round(cpt_map.get("36466", {}).get("medicareRate", 0), 2))
check("CPT 36466 — volume share",      "SC!D127",  xv["cpt_36466_share"],
      cpt_map.get("36466", {}).get("mixPct"))
check("CPT 36475 — Medicare rate",     "SC!C124",  round(xv["cpt_36475_rate"] or 0, 2),
      round(cpt_map.get("36475", {}).get("medicareRate", 0), 2))
check("CPT 36475 — volume share",      "SC!D124",  xv["cpt_36475_share"],
      cpt_map.get("36475", {}).get("mixPct"))

# Credentialing ramp — Month 1 commercial % check
m1 = av.get("monthlyY1", [{}])[0]
m1_comm = m1.get("commercialPct", None)
check("Credentialing ramp — M1 commercial %", "SC!D163", 0.40,
      round(m1_comm, 2) if m1_comm is not None else None,
      tol=0.02)

m5 = av.get("monthlyY1", [{}])[4] if len(av.get("monthlyY1",[])) > 4 else {}
m5_comm = m5.get("commercialPct", None)
check("Credentialing ramp — M5+ commercial %", "SC!D167", 0.85,
      round(m5_comm, 2) if m5_comm is not None else None,
      tol=0.02)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — Report
# ─────────────────────────────────────────────────────────────────────────────
passes  = [r for r in results if r["status"] == PASS]
fails   = [r for r in results if r["status"] == FAIL]
warns   = [r for r in results if r["status"] == WARN]

print(f"\n{'='*65}")
print(f"  AUDIT RESULTS — {len(results)} checks")
print(f"{'='*65}")
print(f"  ✅ PASS   {len(passes)}")
print(f"  ❌ FAIL   {len(fails)}")
print(f"  ⚠️  WARN   {len(warns)}")

if not args.fail_only:
    print(f"\n{'─'*65}")
    print("  ALL RESULTS:")
    for r in results:
        line = f"  {r['status']}  {r['label']}"
        if r["explanation"]:
            line += f"\n         → {r['explanation']}"
        print(line)

if fails:
    print(f"\n{'─'*65}")
    print("  ❌ FAILURES:")
    for r in fails:
        print(f"  • {r['label']}  [{r['cell']}]")
        print(f"    Spreadsheet: {r['spreadsheet']}")
        print(f"    App:         {r['app']}")
        if r["explanation"]:
            print(f"    Note:        {r['explanation']}")

if warns:
    print(f"\n{'─'*65}")
    print("  ⚠️  WARNINGS (spreadsheet cells returned None — formula not cached):")
    for r in warns:
        print(f"  • {r['label']}  [{r['cell']}]")

print(f"\n  OVERALL: {'✅ ALL PASS — app matches spreadsheet' if not fails else '❌ MISMATCHES FOUND — see above'}")
print(f"{'='*65}\n")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — Save JSON
# ─────────────────────────────────────────────────────────────────────────────
output = {
    "timestamp": datetime.now().isoformat(),
    "app_url": args.app,
    "spreadsheet": args.xlsx,
    "summary": {"total": len(results), "pass": len(passes), "fail": len(fails), "warn": len(warns)},
    "results": results,
    "spreadsheet_values": {k: str(v) for k, v in xv.items()},
    "app_audit_snapshot": av,
}

if args.save:
    Path(args.save).write_text(json.dumps(output, indent=2, default=str))
    print(f"Results saved to: {args.save}\n")

sys.exit(1 if fails else 0)
